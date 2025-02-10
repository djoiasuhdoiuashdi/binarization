import pandas as pd
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

def parse_approach(approach):
    """
    Parses the 'Approach' string to extract the Name and parameters.

    Parameters:
        approach (str): The approach string (e.g., 'BERNSEN_window117_threshold190_contrast-limit12').

    Returns:
        tuple: (Name, params_dict)
            - Name (str): The extracted name (e.g., 'BERNSEN').
            - params_dict (dict): Dictionary of parameters and their values.
    """
    if not isinstance(approach, str):
        return None, {}

    parts = approach.split('_')
    name = parts[0]
    params = {}

    # Iterate over the remaining parts to extract parameters
    for part in parts[1:]:
        # Use regex to match parameter patterns with hyphens
        match = re.match(r'^([a-zA-Z\-]+?)(\d+)$', part)
        if match:
            param_name = match.group(1)
            param_value = match.group(2)
            params[param_name] = param_value
        else:
            # Handle unexpected formats or parameters without numerical values
            params[part] = None

    return name, params

def transform_excel(input_file='averages.xlsx', output_dir='output_files', sheet_name='Sheet1'):
    """
    Transforms the Excel data by extracting Name and Parameters from the 'Approach' column
    and creates separate Excel files for each unique Name.

    Parameters:
        input_file (str): Path to the input Excel file.
        output_dir (str): Directory to save the transformed Excel files.
        sheet_name (str): Name of the sheet to read from.
    """
    # Read the Excel file
    try:
        df = pd.read_excel(input_file, sheet_name=sheet_name)
    except FileNotFoundError:
        print(f"Error: The file '{input_file}' was not found.")
        return
    except Exception as e:
        print(f"An error occurred while reading '{input_file}': {e}")
        return

    # Check if 'Approach' column exists
    if 'Approach' not in df.columns:
        print("Error: 'Approach' column not found in the Excel file.")
        return

    # Lists to store parsed data
    names = []
    param_dicts = []

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        approach = row.get('Approach')  # Safely get the 'Approach' column
        if pd.isna(approach):
            print(f"Warning: Missing 'Approach' value at row {index}. Skipping this row.")
            names.append(None)
            param_dicts.append({})
            continue
        name, params = parse_approach(approach)
        names.append(name)
        param_dicts.append(params)

    # Add the parsed 'Name' to the DataFrame
    df['Name'] = names

    # Convert list of parameter dicts to a DataFrame
    params_df = pd.DataFrame(param_dicts)

    # Optionally, convert parameter values to numeric types if applicable
    params_df = params_df.apply(pd.to_numeric, errors='ignore')

    # Identify parameters that already exist in the original DataFrame
    existing_params = set(df.columns) - {'Approach', 'Name'}

    # Update existing parameter columns with parsed values
    for param in existing_params:
        if param in params_df.columns:
            df[param] = params_df[param].combine_first(df[param])

    # Identify new parameters not present in the original DataFrame
    new_params = params_df.columns.difference(existing_params)

    # Add new parameter columns to the DataFrame
    for param in new_params:
        df[param] = params_df[param]

    # Drop the original 'Approach' column
    df.drop(columns=['Approach'], inplace=True)

    # -----------------------------------------
    # **Modified Section to Keep Column Order**
    # -----------------------------------------

    # Store the original column order excluding 'Approach'
    original_columns = list(pd.read_excel(input_file, sheet_name=sheet_name).columns)
    original_columns.remove('Approach')  # Remove 'Approach' as it will be replaced

    # Determine the position where 'Approach' was located
    approach_index = len(original_columns)  # Default to end if not found

    original_df = pd.read_excel(input_file, sheet_name=sheet_name)
    approach_index = original_df.columns.get_loc('Approach')

    # Prepare the list of new columns to insert
    # 'Name' followed by sorted new parameters
    insert_columns = ['Name'] + sorted(new_params.tolist())

    # Prepare the final list of columns
    # Insert the new columns at the position of 'Approach'
    final_columns = (
        original_columns[:approach_index] +
        insert_columns +
        original_columns[approach_index:]
    )

    # Add any additional columns that might have been added outside original_columns
    additional_columns = set(df.columns) - set(final_columns)
    final_columns += sorted(additional_columns)

    # Reorder the DataFrame columns to match the final_columns
    transformed_df = df.reindex(columns=final_columns)

    # Handle missing parameters by filling with NaN or a default value
    transformed_df.fillna(value=pd.NA, inplace=True)

    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Group the DataFrame by 'Name' and save each group to a separate Excel file
    grouped = transformed_df.groupby('Name')
    
    for name, group in grouped:
        if pd.isna(name):
            print(f"Warning: Encountered a group with 'Name' as NaN. Skipping this group.")
            continue

        # Define a safe filename by removing or replacing invalid characters
        safe_name = re.sub(r'[\\/*?:\"<>|]', "_", str(name))
        output_file = os.path.join(output_dir, f"{safe_name}.xlsx")

        # Identify parameter columns used in this group (non-NaN)
        used_params = group.drop(columns=['Name']).loc[:, group.drop(columns=['Name']).notna().any(axis=0)].columns.tolist()

        # Include 'Name' and used parameters only
        columns_to_include = ['Name'] + used_params

        # Create a DataFrame with only the relevant columns
        group_filtered = group[columns_to_include]

        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                group_filtered.to_excel(writer, index=False, sheet_name='Sheet1')

                worksheet = writer.sheets['Sheet1']
                
                # Set column width to 25
                for i, col in enumerate(group_filtered.columns, 1):
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = 10
                
                # Define a thick border
                thick = Side(border_style="thick", color="000000")
                border = Border(right=thick)

                # Add thick lines between ID ranges (this part needs to be adjusted based on your logic)
                headers = group_filtered.columns.tolist()

                current_id = None
                for idx, col in enumerate(headers, 1):
                    if col == 'Approach' or col == "contrast-limit" or col == "window" or col =="threshold" or col =="k" or col =="glyph" or col =="minN":
                        continue
                    id_range = col.split('_')[-1]
                    if current_id and id_range != current_id:
                        col_letter = get_column_letter(idx-1)
                        for row in range(1, len(df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.border = cell.border + border
                    current_id = id_range

            print(f"Saved '{safe_name}.xlsx' with {len(group_filtered)} records and {len(used_params)} parameters.")
        except Exception as e:
            print(f"Failed to save '{safe_name}.xlsx': {e}")

    print(f"\nAll transformed files have been saved to the '{output_dir}' directory.")

if __name__ == "__main__":
    transform_excel()