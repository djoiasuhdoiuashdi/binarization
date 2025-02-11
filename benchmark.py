import os
from pathlib import Path
import numpy as np
from concurrent.futures import ProcessPoolExecutor, as_completed
import logging
from functools import partial
from tqdm import tqdm
from metrics import load_image_as_binary, calculate_metrics
from collections import defaultdict
import pandas as pd
from colorama import init, Fore, Style
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

init(autoreset=True)

def get_image_files(input_dir):
    input_path = Path(input_dir)
    image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'}
    return [file for file in input_path.rglob('*') if file.suffix.lower() in image_extensions]


def process_file(file_path, gt_path_a, gt_path_b, weight_a, weight_b):
    try:
        # Load the input image
        im = load_image_as_binary(file_path)

        gt_name = file_path.stem
        gt_num = int(gt_name)
        if gt_num <= 10:
            gt_dir = gt_path_a
            weight_dir = weight_a
        else:
            gt_dir = gt_path_b
            weight_dir = weight_b

        gt_file = gt_dir / f"{gt_name}.bmp"
        im_gt = load_image_as_binary(gt_file)

        height, width = im_gt.shape
        r_weights_file = weight_dir / f"{gt_name}_RWeights.dat"
        p_weights_file = weight_dir / f"{gt_name}_PWeights.dat"

        r_weight = np.loadtxt(r_weights_file, dtype=np.float64).flatten()[:height * width].reshape((height, width))
        p_weight = np.loadtxt(p_weights_file, dtype=np.float64).flatten()[:height * width].reshape((height, width))

        f_measure, w_f_measure, psnr, drd, recall, precision, w_recall, w_precision = calculate_metrics(im, im_gt,
                                                                                                        r_weight,
                                                                                                        p_weight)

        f_measure *= 100
        w_f_measure *= 100
        recall *= 100
        precision *= 100
        w_recall *= 100
        w_precision *= 100

        result = {
            "id": gt_num,
            "subdir": file_path.parent.name,
            "FM": f_measure,
            "pFM": w_f_measure,
            "PSNR": psnr,
            "DRD": drd,
            "Rec": recall,
            "Pre": precision,
            "pRec": w_recall,
            "pPre": w_precision
        }
        return file_path, None, result
    except Exception as e:
        error_msg = f"Failed to process {file_path}: {e}"
        return file_path, e, error_msg


def group_metrics_by_subdir_and_id_range(metrics):
    grouped = defaultdict(lambda: defaultdict(list))
    for metric in metrics:
        subdir = metric["subdir"]
        grouped[subdir]['ALL'].append(metric)
        id_num = metric["id"]
        id_range = 'A' if 1 <= id_num <= 10 else 'B'
        grouped[subdir][id_range].append(metric)
    return grouped


def calculate_averages(grouped_metrics, metric_keys):
    averages = {}
    for subdir, id_groups in grouped_metrics.items():
        averages[subdir] = {}
        for id_range, metrics in id_groups.items():
            if metrics:
                avg_metrics = {
                    key: sum(m[key] for m in metrics) / len(metrics)
                    for key in metric_keys
                }
                averages[subdir][id_range] = avg_metrics
            else:
                averages[subdir][id_range] = {key: None for key in metric_keys}
    return averages


def save_averages_to_excel(averages, filename=None):
    now = datetime.now()
    formatted_time = now.strftime("%y-%m-%d-%H-%M")
    if filename is None:
        filename = f"{formatted_time}-Results.xlsx"

    data = []
    for subdir, id_ranges in averages.items():
        row = {'Approach': subdir}
        for id_range, metrics in id_ranges.items():
            for key, value in metrics.items():
                column_name = f"{key}_{id_range}"
                row[column_name] = value
        data.append(row)

    df = pd.DataFrame(data)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        
        # Set column width to 25
        for i, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = 10
        
        # Define thick border
        thick = Side(border_style="thick", color="000000")
        border = Border(right=thick)
        
        # Add thick lines between id_ranges
        headers = df.columns.tolist()

        current_id = None
        for idx, col in enumerate(headers, 1):
            if col == 'Approach':
                continue
            id_range = col.split('_')[-1]
            if current_id and id_range != current_id:
                col_letter = get_column_letter(idx - 1)
                for row in range(1, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.border = cell.border + border
            current_id = id_range

        numeric_columns = [col for col in df.columns if col != 'Approach']

        for col in numeric_columns:
            col_idx = df.columns.get_loc(col) + 1
            col_letter = get_column_letter(col_idx)

            for row in range(2, len(df) + 2):
                cell = worksheet[f"{col_letter}{row}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'  # Adjust as needed


def print_averages(averages):
    for subdir, id_ranges in averages.items():
        print(Fore.CYAN + f"Approach: {subdir}")
        for id_range, metrics in id_ranges.items():
            print(Fore.GREEN + f"  ID Range: {id_range}")
            for key, value in metrics.items():
                print(Fore.YELLOW + f"    {key}: " + Fore.RED + f"{value:.4f}")

        print(Style.RESET_ALL + "\n")


def main():
    input_dir = "./benchmark_input"
    gt_path_a = Path("./ground_truth/trackA-GT")
    gt_path_b = Path("./ground_truth/trackB-GT")
    weight_a = Path("./ground_truth/trackA-weights")
    weight_b = Path("./ground_truth/trackB-weights")

    image_files = get_image_files(input_dir)
    logging.info(f"Found {len(image_files)} image files to process.")

    max_workers = os.cpu_count()

    process_partial = partial(
        process_file,
        gt_path_a=gt_path_a,
        gt_path_b=gt_path_b,
        weight_a=weight_a,
        weight_b=weight_b
    )
    # List to collect successful metrics
    metrics_list = []

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_partial, file_path): file_path for file_path in image_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Processing images"):
            file_path = futures[future]
            try:
                processed_file, error, message = future.result()
                if error:
                    logging.error(message)
                else:
                    metrics_list.append(message)
            except Exception as e:
                logging.error(f"Unhandled exception for {file_path}: {e}")

    grouped_metrics = defaultdict(list)
    for metric in metrics_list:
        grouped_metrics[metric["subdir"]].append(metric)

    metric_keys = [
        "FM",
        "pFM",
        "PSNR",
        "DRD",
        "Rec",
        "Pre",
        "pRec",
        "pPre"
    ]

    grouped = group_metrics_by_subdir_and_id_range(metrics_list)
    averages = calculate_averages(grouped, metric_keys)
    print_averages(averages)
    save_averages_to_excel(averages)


if __name__ == '__main__':
    main()